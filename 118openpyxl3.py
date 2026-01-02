from openpyxl import load_workbook

wb = load_workbook("AlunosNotas.xlsx")
ws = wb.active

for linha in ws.iter_rows(min_row=2, values_only=True):
    nome, nota = linha
    print(nome, nota)